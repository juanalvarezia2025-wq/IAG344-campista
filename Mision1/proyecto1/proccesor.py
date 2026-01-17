import re # Importa el módulo de expresiones regulares
from openpyxl import load_workbook # Importa la función load_workbook desde la librería openpyxl para trabajar con archivos Excel

# ============================= 
#  🤖 PUNTO DE ENTRADA
# Función para eliminar catecteres no numericos de un documento.
# ============================= 

def clear_id(value):
    
    if value is None:
        return ""
    return re.sub(r'\D', '', str(value)) # Reemplaza todos los caracteres no numéricos con una cadena vacía, sub es una función de re que realiza esta operación

# ============================= 
#  🤖 FUNCION marge_name
# Une Nombre y apellidos en un solo campo
# ============================= 

def merge_name(name, last_name):
    if name is None:
        name = ""
    if last_name is None:
        last_name = ""
    return f"{name} {last_name}".strip()

# ============================= 
#  🤖 
# ============================= 

def process_excel(path):
    # Importa la librería openpyxl para trabajar con archivos Excel - Acceso a la hoja llada "Datos"
    wb = load_workbook(path)
    ws = wb["Datos"]
    # Recorre las filas de la hoja de Excel
    for row in range (2,ws.max_row + 1):
       # Columna D = ID limpio
       ws[f"D{row}"] = clear_id(ws[f"A{row}"].value) # Limpia el ID en la columna A y lo guarda en la columna D
       
       # Columna E = Nombre completo
       ws[f"E{row}"] = merge_name(
       ws[f"B{row}"].value, 
       ws[f"C{row}"].value
       ) # Une el nombre y apellido en las columnas B y C y lo guarda en la columna E

    wb.save(path) # Guarda los cambios en el archivo Excel

def proccess_excel_safe(path):
    try:
        process_excel(path)
        return True, "Procesado exitosamente"
    except PermissionError:
        return (
            False, 
            "Error de permiso: No se puede acceder al archivo.\n" 
            "Por favor, cierre el archivo si está abierto e intente nuevamente."
        )
    except KeyError:
        return (
            False, 
            "Error de clave: La hoja 'Datos' no se encontró en el archivo.\n" 
            "Por favor, verifique el archivo e intente nuevamente."
        )
    except Exception as e:
        return (
            False, 
            f"Error inesperado: {str(e)}\n" 
            "Por favor, verifique el archivo e intente nuevamente."
        )